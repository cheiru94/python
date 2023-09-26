import openpyxl

fpath = r'C:\Python_File\4. 인프런\03. 파이썬 엑셀 다루기\참가자_data.xlsx'

# 1) 엑셀 불러오기 => .load_workbook('파일이름')
wb = openpyxl.load_workbook(fpath)

# 2) 엑셀 시트선택
ws = wb['우리집']

# 3) 데이터 수정하기
ws['A3'] = 7068
ws['B3'] = '이재일'

# 4) 엑셀 저장하기 
wb.save(fpath)
