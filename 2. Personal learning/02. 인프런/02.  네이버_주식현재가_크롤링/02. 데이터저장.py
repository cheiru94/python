import requests
from bs4 import BeautifulSoup
import openpyxl

fpath = r'C:\Python_File\4. 인프런\02.  네이버_주식현재가_크롤링\data.xlsx'

# 엑셀 불러오기 .load_workbook()
wb = openpyxl.load_workbook(fpath)

#  .active 현재 활성화된 시트 선택 (시트 한개 밖에 없을 때 기본시트 선택 )
ws = wb.active   # 


# 종목 코드 리스트
Codes = [
    '005930',
    '000660',
    '035720'
]


row = 2
for Code in Codes : 
    url = f"https://finance.naver.com/item/sise.naver?code={Code}"

    # requests 라이브러리에 get 함수를 사용해서 사이트 요청 하기 (해당 url사이트로 get 요청이감)
    # ->반환된 응답을 "response" 변수에 저장하기    
    response = requests.get(url)

    # 응답 받은 내용 html 코드 문서로 만들기(. text를 사용해 문자 내용만 가져오기)
    # 받은 html 코드는 문자이기 떄문에 파씽하기가 어려움
    html = response.text

    # 때문에 받은 html 문서에 속성값 인지 할 수 있게 BeautifulSoup로 만들기
    # BeautifulSoup(첫번째 인자 = response.text , 두번째 인자 = "html.parser" )     
    soup = BeautifulSoup(html,"html.parser")

    # 문자열 형태로 CSS 선택자를 써준다 
    price = soup.select_one("#_nowVal").text 

    # replace 함수(문자열 교체함수)를 사용해서 원하는 문자열 변경    
    # =>    .replace( " 변경 할 문자 " , " 변경되는 문자 ")
    price = price.replace(",","")
    # 출력 
    print(price)


    # 시트(셀) 에 추가로 입력    -> 값 int 로 변경 
    ws[f'B{row}'] = int(price)
    # 셀 변강하기 위해 1씩 증가 
    row += 1
# 파일 저장   -> .save 
wb.save(fpath)